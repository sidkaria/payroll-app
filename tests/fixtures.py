"""Synthetic fixture builders for payroll tests.

These build ADP CSVs and notes XLSX files in temp directories using only
generic placeholder names (e.g. "Doe, Jane") — never real employee data.
Keeps the test suite hermetic and PII-free.
"""

from __future__ import annotations

import csv
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


# ── ADP CSV builder ──────────────────────────────────────────────────────────

@dataclass
class AdpEntry:
    employee: str
    date: str          # "M/D/YYYY"
    in_time: str       # "7:00 AM"
    out_time: str      # "4:00 PM"
    hours: float


def write_adp_csv(path: Path, pay_start: str, pay_end: str, entries: list[AdpEntry]) -> Path:
    """Write a minimal ADP CSV matching the columns the parser reads."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Date range", pay_start, pay_end])
        writer.writerow([
            "Employee Name", "Pay number", "Date", "In time", "Out time",
            "Hours", "Position", "Department", "Job", "Note",
        ])
        last_emp = None
        for e in entries:
            name_cell = e.employee if e.employee != last_emp else ""
            last_emp = e.employee
            writer.writerow([
                name_cell, "", e.date, e.in_time, e.out_time,
                f"{e.hours:.2f}", "", "", "", "",
            ])
    return path


# ── Notes XLSX builder ───────────────────────────────────────────────────────

@dataclass
class NoteRow:
    name: str
    status: str = "Full Time"
    schedule: str = ""           # e.g. "9:00 am - 6:00 pm"
    break_time: str = ""         # e.g. "1 hour break"
    missed_date: str = ""
    missed_punches: str = ""
    outside_date: str = ""
    outside_desc: str = ""
    time_off_date: str = ""      # currently unused — left blank, lump-sum
    time_off_hours: str = ""
    holiday_date: str = ""
    holiday_hours: str = ""
    notes: str = ""


@dataclass
class NotesWorkbook:
    rows: list[NoteRow] = field(default_factory=list)
    prompts: list[str] = field(default_factory=list)
    sheet_name: str = "Sheet1"


def write_notes_xlsx(path: Path, workbook_data: NotesWorkbook) -> Path:
    """Build a notes workbook with the same layout the parser expects."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = workbook_data.sheet_name

    # Row 4: column headers (the parser searches col 1 for "Employee")
    ws.cell(4, 1, "Employee")
    ws.cell(4, 2, "Status")
    ws.cell(4, 3, "Scheduled hours")
    ws.cell(4, 4, "Break Time")
    ws.cell(4, 5, "Missed punches")
    ws.cell(4, 7, "Worked Outside Scheduled Hours")
    ws.cell(4, 9, "Time Off")
    ws.cell(4, 11, "Holiday")
    ws.cell(4, 13, "Notes")

    # Row 5: sub-headers
    ws.cell(5, 5, "Date")
    ws.cell(5, 6, "Clock In/Out:")
    ws.cell(5, 7, "Date")
    ws.cell(5, 8, "Description")
    ws.cell(5, 9, "Date")
    ws.cell(5, 10, "Hours")
    ws.cell(5, 11, "Date")
    ws.cell(5, 12, "Hours")

    for index, row in enumerate(workbook_data.rows):
        r = 6 + index
        ws.cell(r, 1, row.name)
        ws.cell(r, 2, row.status)
        ws.cell(r, 3, row.schedule)
        ws.cell(r, 4, row.break_time)
        ws.cell(r, 5, row.missed_date)
        ws.cell(r, 6, row.missed_punches)
        ws.cell(r, 7, row.outside_date)
        ws.cell(r, 8, row.outside_desc)
        ws.cell(r, 9, row.time_off_date)
        ws.cell(r, 10, row.time_off_hours)
        ws.cell(r, 11, row.holiday_date)
        ws.cell(r, 12, row.holiday_hours)
        ws.cell(r, 13, row.notes)

    # Optional prompts section
    if workbook_data.prompts:
        prompt_row = 6 + len(workbook_data.rows) + 5
        ws.cell(prompt_row, 1, "Prompts:")
        for offset, text in enumerate(workbook_data.prompts, start=1):
            ws.cell(prompt_row + offset, 1, text)

    wb.save(path)
    return path


# ── Quick path helper ────────────────────────────────────────────────────────

def temp_path(suffix: str) -> Path:
    """Allocate a temp file path. Caller is responsible for cleanup."""
    fd, name = tempfile.mkstemp(suffix=suffix)
    Path(name).unlink()  # we just want the path; openpyxl/csv will create the file
    import os
    os.close(fd)
    return Path(name)
