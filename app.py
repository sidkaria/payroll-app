import os
import tempfile
import traceback
from collections import defaultdict
from io import BytesIO
from pathlib import Path

import streamlit as st
import openpyxl

from payroll_calculator import (
    apply_notes,
    fmt_hours,
    normalize_name,
    parse_adp,
    parse_notes,
    parse_workbook_rules,
    weekday_dates_only,
    write_excel,
)

# ── Page config ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Bel Air Payroll Calculator",
    page_icon="📋",
    layout="centered",
)

st.title("📋 Bel Air Payroll Calculator")
st.caption(
    "Upload your ADP export (CSV) and payroll notes (Excel) to generate a corrected hours report."
)
st.divider()

# ── File uploaders ────────────────────────────────────────────────────────────

col_a, col_b = st.columns(2)

with col_a:
    adp_file = st.file_uploader(
        "**ADP Export**",
        type=["csv", "xlsx"],
        help="The CSV exported from ADP — e.g. ADPPayroll.csv. File name doesn't matter.",
    )

with col_b:
    notes_file = st.file_uploader(
        "**Payroll Notes**",
        type=["xlsx", "xlsm", "docx"],
        help="Your payroll notes Excel sheet with schedules, missed punches, and exceptions. File name doesn't matter.",
    )

# ── Validation ────────────────────────────────────────────────────────────────

def validate_adp(raw: bytes) -> tuple[bool, str]:
    try:
        first_line = raw.decode("utf-8-sig").split("\n")[0]
        if "Date range" in first_line or "Employee Name" in first_line:
            return True, ""
        return False, "First row doesn't look like an ADP export. Expected 'Date range' or 'Employee Name' header."
    except Exception as exc:
        return False, str(exc)


def validate_notes_xlsx(raw: bytes) -> tuple[bool, str]:
    try:
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            text = " ".join(str(v or "") for v in row)
            if "Scheduled hours" in text or "Payroll" in text or "Schedule" in text:
                return True, ""
        return False, "This Excel doesn't look like the payroll notes sheet — expected a 'Scheduled hours' column. Proceeding anyway, but check your output."
    except Exception as exc:
        return False, str(exc)


adp_ok = notes_ok = True
adp_bytes = notes_bytes = None

if adp_file:
    adp_bytes = adp_file.read()
    adp_file.seek(0)
    adp_ok, adp_msg = validate_adp(adp_bytes)
    if not adp_ok:
        st.error(f"ADP file issue: {adp_msg}")

if notes_file:
    notes_bytes = notes_file.read()
    notes_file.seek(0)
    ext = Path(notes_file.name).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        notes_ok, notes_msg = validate_notes_xlsx(notes_bytes)
        if not notes_ok:
            st.warning(f"Notes file warning: {notes_msg}")
            notes_ok = True  # still allow run, just warn

# ── Run button ────────────────────────────────────────────────────────────────

both_uploaded = adp_file is not None and notes_file is not None
st.divider()

if not both_uploaded:
    st.info("Upload both files above to enable the calculator.")

run = st.button(
    "▶  Calculate Payroll",
    type="primary",
    disabled=not (both_uploaded and adp_ok),
    use_container_width=True,
)

# ── Calculation ───────────────────────────────────────────────────────────────

if run:
    with st.spinner("Calculating…"):
        try:
            # Save uploads to temp files (the calculator expects file paths)
            ext_adp = Path(adp_file.name).suffix.lower()
            ext_notes = Path(notes_file.name).suffix.lower()

            with tempfile.NamedTemporaryFile(suffix=ext_adp, delete=False) as f:
                adp_file.seek(0)
                f.write(adp_file.read())
                tmp_csv = f.name

            with tempfile.NamedTemporaryFile(suffix=ext_notes, delete=False) as f:
                notes_file.seek(0)
                f.write(notes_file.read())
                tmp_notes = f.name

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                tmp_out = f.name

            # Core calculation
            adp = parse_adp(tmp_csv)
            entries = adp["entries"]
            notes = parse_notes(tmp_notes)
            rules = parse_workbook_rules(notes)
            day_records, logs, leave_totals = apply_notes(entries, notes, rules)

            pay_start = adp["pay_period_start"]
            pay_end = adp["pay_period_end"]
            if not pay_start or not pay_end:
                dates = sorted(r["date"] for r in day_records.values())
                pay_start, pay_end = dates[0], dates[-1]

            pay_dates = weekday_dates_only(pay_start, pay_end)
            employee_meta = notes.get("employees", {})

            write_excel(
                tmp_out, day_records, logs, leave_totals,
                employee_meta, pay_dates, pay_start, pay_end,
            )

            with open(tmp_out, "rb") as f:
                output_bytes = f.read()

        except Exception as exc:
            st.error(f"Something went wrong: {exc}")
            with st.expander("Error details"):
                st.code(traceback.format_exc())
            output_bytes = None
        finally:
            for p in [tmp_csv, tmp_notes, tmp_out]:
                try:
                    os.unlink(p)
                except Exception:
                    pass

    if output_bytes:

        # ── Summary metrics ───────────────────────────────────────────────────

        st.divider()
        st.subheader("Results")

        all_employees = sorted(
            set(employee_meta.keys()) | {r["employee"] for r in day_records.values()},
            key=normalize_name,
        )

        total_worked = sum(r["final_work_hours"] for r in day_records.values())
        total_leave = sum(
            leave_totals[e]["Time Off"] + leave_totals[e]["Holiday"]
            for e in leave_totals
        )
        total_payable = round(total_worked + total_leave, 2)

        n_corrected = sum(1 for c in logs["corrections"] if c["status"] == "CORRECTED")
        n_review = len(logs["review"]) + sum(
            1 for c in logs["corrections"] if c["status"] == "REVIEW"
        )

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Employees", len(all_employees))
        m2.metric("Total Payable", fmt_hours(total_payable))
        m3.metric("Corrections", n_corrected)
        if n_review:
            m4.metric("⚠ Needs Review", n_review)
        else:
            m4.metric("Needs Review", 0)

        # ── Corrections ───────────────────────────────────────────────────────

        if logs["corrections"]:
            st.subheader("Punch Corrections")
            for item in logs["corrections"]:
                date_str = item["date"].strftime("%m/%d/%Y")
                if item["status"] == "CORRECTED":
                    delta = round((item["corrected_hours"] or 0) - item["raw_hours"], 2)
                    st.success(
                        f"✓ **{item['employee']}** · {date_str} · "
                        f"{item['raw_hours']:.2f} → **{item['corrected_hours']:.2f} hrs** "
                        f"({delta:+.2f})  —  {item['note']}"
                    )
                else:
                    st.warning(
                        f"⚠ **{item['employee']}** · {date_str} · "
                        f"{item.get('note', 'needs review')}"
                    )

        # ── Review flags ──────────────────────────────────────────────────────

        if logs["review"]:
            st.subheader("Needs Review")
            st.caption(
                "These items were unclear or could not be calculated automatically. "
                "Check each one and correct the notes sheet if needed."
            )
            for item in logs["review"]:
                emp = item.get("employee", "Unknown")
                date_val = item.get("date", "")
                date_str = date_val.strftime("%m/%d/%Y") if hasattr(date_val, "strftime") else str(date_val)
                issue = item.get("issue", "unknown issue")
                st.error(f"⚠ **{emp}** · {date_str} · {issue}")

        # ── Approved exceptions ───────────────────────────────────────────────

        if logs.get("exceptions"):
            with st.expander(f"Approved schedule exceptions ({len(logs['exceptions'])})"):
                for item in logs["exceptions"]:
                    st.write(
                        f"• **{item['employee']}** · "
                        f"{item['date'].strftime('%m/%d/%Y')} · "
                        f"{item['description']}"
                    )

        # ── Download ──────────────────────────────────────────────────────────

        st.divider()
        out_name = f"Payroll_{pay_start.strftime('%m%d%y')}_{pay_end.strftime('%m%d%y')}_corrected.xlsx"
        st.download_button(
            label="📥  Download Corrected Excel",
            data=output_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
