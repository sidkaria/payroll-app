#!/usr/bin/env python3
"""
Payroll Hours Calculator

Processes ADP CSV time data plus either:
- the original School-based Report DOCX, or
- Megha's newer payroll notes XLSX workbook.

The XLSX workflow is deterministic and day-based:
- missed punches override ADP hours when the note is explicit
- worked hours are capped to the scheduled day unless an approved exception exists
- holiday and time-off hours are added as payroll adjustments
- one-off prompt rules from the workbook are applied when present

Usage:
    python3 payroll_calculator.py <adp_csv> <notes.docx|notes.xlsx>
"""

import csv
import re
import sys
from collections import defaultdict
from datetime import date, datetime, time as dtime, timedelta
from pathlib import Path

import docx
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# TIME HELPERS
# ---------------------------------------------------------------------------

TIME_TOKEN_RE = re.compile(
    r"\d{1,2}:\d{2}\s*(?:[AaPp][Mm])?|\d{1,2}\s*(?:[AaPp][Mm])"
)
DATE_RE = re.compile(r"\d{1,2}/\d{1,2}(?:/\d{2,4})?")


def parse_time(value):
    """
    Parse a time string into datetime.time.

    Handles:
    - 9:30 am / 9:30am
    - 9 am
    - 14:30
    - bare times like 4:30 using a payroll-friendly heuristic
    """
    if value is None:
        return None

    # Already a time object (e.g. from openpyxl reading an Excel ADP export)
    if isinstance(value, dtime):
        return value

    text = str(value).strip().rstrip(".")
    if not text:
        return None

    for fmt in ("%I:%M %p", "%I:%M%p", "%I %p", "%I%p"):
        try:
            return datetime.strptime(text.upper(), fmt).time()
        except ValueError:
            pass

    try:
        return datetime.strptime(text, "%H:%M").time()
    except ValueError:
        pass

    m = re.fullmatch(r"(\d{1,2})(?::(\d{2}))?", text)
    if not m:
        return None

    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    try:
        if 1 <= hour <= 5:
            return dtime(hour + 12, minute)
        if hour == 12:
            return dtime(12, minute)
        return dtime(hour, minute)
    except ValueError:
        return None


def make_after(value, reference):
    """Flip a bare earlier time into the afternoon when needed."""
    if value is None or reference is None:
        return value

    base = datetime(2000, 1, 1)
    if datetime.combine(base, value) < datetime.combine(base, reference):
        hour = (value.hour + 12) % 24
        try:
            return dtime(hour, value.minute)
        except ValueError:
            return value
    return value


def hours_between(start, end):
    if start is None or end is None:
        return 0.0
    base = datetime(2000, 1, 1)
    return (datetime.combine(base, end) - datetime.combine(base, start)).total_seconds() / 3600


def fmt_hours(hours):
    if hours is None:
        return ""
    total_minutes = round(abs(hours) * 60)
    sign = "-" if hours < 0 else ""
    return f"{sign}{total_minutes // 60}:{total_minutes % 60:02d}"


def daterange(start_date, end_date):
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


# ---------------------------------------------------------------------------
# GENERAL HELPERS
# ---------------------------------------------------------------------------

ROLE_PREFIX_RE = re.compile(r"^(SubAide|Yoga|Bel Air)\s*", re.I)


def normalize_name(name):
    text = ROLE_PREFIX_RE.sub("", str(name or "").strip())
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        text = f"{first} {last}"
    return re.sub(r"\s+", " ", text).strip().lower()


def resolve_employee_name(name, known_names):
    text = str(name or "").strip()
    if not text:
        return None, "missing employee name"

    if text in known_names:
        return text, None

    normalized = normalize_name(text)
    matches = [candidate for candidate in known_names if normalize_name(candidate) == normalized]
    if len(matches) == 1:
        return matches[0], None
    if len(matches) > 1:
        return None, f"ambiguous match for {text}: {', '.join(sorted(matches))}"
    return None, f"no ADP match for {text}"


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return str(value).strip()


def split_nonempty_lines(value):
    text = cell_to_text(value)
    return [line.strip() for line in text.splitlines() if line.strip()]


def parse_float(value):
    if value is None:
        return None
    # openpyxl returns datetime.time for cells like "8:00" — treat as hours
    if isinstance(value, dtime):
        return round(value.hour + value.minute / 60, 4)
    text = str(value).strip()
    if not text:
        return None
    # Multiline cells (e.g. "8:00\n8:00\n8:00") — sum each line
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) > 1:
        total = sum(v for ln in lines for v in [parse_float(ln)] if v is not None)
        return round(total, 4) if total > 0 else None
    # "H:MM" format like "8:00" or "2:30"
    m = re.match(r'^(\d+):(\d{2})$', text)
    if m:
        return round(int(m.group(1)) + int(m.group(2)) / 60, 4)
    # "8H" or "8 hrs" etc.
    m2 = re.match(r'^(\d+(?:\.\d+)?)\s*[hH]', text)
    if m2:
        return float(m2.group(1))
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def parse_date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    # Year-less form like "4/8" — default to the current year
    try:
        parsed = datetime.strptime(text, "%m/%d").date()
        return parsed.replace(year=date.today().year)
    except ValueError:
        return None


def expand_date_expression(value):
    """
    Parse date cells like:
    - 03/18/26
    - 03/09/26 03/18/26
    - 03/09/26-03/13/26
    - multiline combinations of the above
    """
    if isinstance(value, (datetime, date)):
        parsed = parse_date_value(value)
        return [parsed] if parsed else []

    text = cell_to_text(value)
    if not text:
        return []

    results = []
    for line in split_nonempty_lines(text):
        line_matches = re.finditer(
            r"(\d{1,2}/\d{1,2}(?:/\d{2,4})?)(?:\s*-\s*(\d{1,2}/\d{1,2}(?:/\d{2,4})?))?",
            line,
        )
        for match in line_matches:
            start = parse_date_value(match.group(1))
            end = parse_date_value(match.group(2)) if match.group(2) else start
            if not start:
                continue
            if not end:
                end = start
            if end < start:
                start, end = end, start
            for current in daterange(start, end):
                results.append(current)

    seen = set()
    ordered = []
    for item in results:
        if item and item not in seen:
            seen.add(item)
            ordered.append(item)
    return ordered


def weekday_dates_only(start_date, end_date):
    if not start_date or not end_date:
        return []
    return [current for current in daterange(start_date, end_date) if current.weekday() < 5]


# ---------------------------------------------------------------------------
# SCHEDULE PARSING
# ---------------------------------------------------------------------------

WEEKDAY_MAP = {
    "M": [0],
    "T": [1],
    "W": [2],
    "TH": [3],
    "F": [4],
    "MWF": [0, 2, 4],
    "TTH": [1, 3],
}


def parse_break_hours(text):
    lowered = text.lower().strip()
    if not lowered or lowered == "-":
        return 0.0
    if "no break" in lowered:
        return 0.0
    # "1/2 hour" = 30 min break
    if re.search(r"1/2\s*hours?", lowered):
        return 0.5
    # "30 min", "30 minutes", "30 minute break"
    m = re.search(r"(\d+)\s*min(?:utes?)?", lowered)
    if m:
        return round(int(m.group(1)) / 60, 4)
    # "deduct 1 hour", "1 hour break", "1 hour", etc.
    match = re.search(r"(?:deduct\s+)?(\d+(?:\.\d+)?)\s*hours?", lowered)
    if match:
        return float(match.group(1))
    return 0.0


def parse_time_range(text):
    matches = list(TIME_TOKEN_RE.finditer(text or ""))
    if len(matches) < 2:
        return None

    start = parse_time(matches[0].group(0))
    end = parse_time(matches[1].group(0))
    if not start or not end:
        return None

    end = make_after(end, start)
    duration = hours_between(start, end)
    if duration <= 0:
        return None

    return {"start": start, "end": end, "duration": duration}


def schedule_info_from_range(start, end, break_hours=0.0, source_text=""):
    duration = hours_between(start, end)
    paid_hours = round(max(0.0, duration - break_hours), 2)
    return {
        "mode": "scheduled",
        "start": start,
        "end": end,
        "break_hours": break_hours,
        "paid_hours": paid_hours,
        "label": source_text.strip(),
    }


def parse_weekday_group(text):
    compact = re.sub(r"\s+", "", text.upper())
    if compact in WEEKDAY_MAP:
        return WEEKDAY_MAP[compact]
    if compact == "TTH":
        return [1, 3]
    if compact == "MWF":
        return [0, 2, 4]
    return []


def parse_schedule_text(text, override_break=None):
    """
    Parse schedule text into a schedule dict.
    override_break: if provided (from a dedicated Break Time column), use it
                    instead of parsing break hours from the schedule text.
    """
    raw = cell_to_text(text)
    if not raw or raw.lower() in ("give worked hours", "as needed"):
        return {"mode": "actual", "label": raw}

    if "|" in raw:
        by_weekday = {}
        for part in [piece.strip() for piece in raw.split("|") if piece.strip()]:
            match = re.match(r"(.+?)\s*-\s*(.+)", part)
            if not match:
                continue
            days = parse_weekday_group(match.group(1))
            timerange = parse_time_range(match.group(2))
            if not days or not timerange:
                continue
            break_hours = override_break if override_break is not None else parse_break_hours(part)
            info = schedule_info_from_range(
                timerange["start"], timerange["end"], break_hours=break_hours, source_text=part
            )
            for weekday in days:
                by_weekday[weekday] = info

        if by_weekday:
            return {"mode": "scheduled", "label": raw, "by_weekday": by_weekday}
        return {"mode": "actual", "label": raw}

    timerange = parse_time_range(raw)
    if timerange:
        break_hours = override_break if override_break is not None else parse_break_hours(raw)
        info = schedule_info_from_range(
            timerange["start"], timerange["end"], break_hours=break_hours, source_text=raw
        )
        by_weekday = {weekday: info for weekday in range(5)}
        return {"mode": "scheduled", "label": raw, "by_weekday": by_weekday}

    return {"mode": "actual", "label": raw}


def schedule_for_date(schedule, current_date):
    if not schedule or schedule.get("mode") != "scheduled":
        return None
    return schedule.get("by_weekday", {}).get(current_date.weekday())


def parse_schedule_override(description, fallback_break=0.0):
    text = cell_to_text(description)
    timerange = parse_time_range(text)
    if not timerange:
        return None

    break_hours = 0.0 if "skip break" in text.lower() else fallback_break
    return schedule_info_from_range(
        timerange["start"], timerange["end"], break_hours=break_hours, source_text=text
    )


# ---------------------------------------------------------------------------
# EXCEPTION TEXT PARSING
# ---------------------------------------------------------------------------

END_TIME_PHRASES = re.compile(
    r"late\s*pickup|pay(?:ed|id)?\s*until|paid\s*until|stayed\s*(?:until|till)|"
    r"worked\s*(?:until|till)|pickup\s*at|picked\s*up|left\s*at|clocked\s*out\s*at",
    re.I,
)
START_TIME_PHRASES = re.compile(
    r"came\s*in\s*(?:early|at)|started\s*(?:early|at)|clocked\s*in\s*at|in\s*at|"
    r"arrived\s*at|early\s*(?:start|arrival)",
    re.I,
)
BARE_RANGE_RE = re.compile(r"(?<!\d)(\d{1,2})\s*(?:[-–]|to)\s*(\d{1,2})(?!\d)", re.I)


def _extract_time_range(text):
    """
    Try to find two times forming a range. First strict (HH:MM or HH am/pm),
    then fall back to bare N-M pattern (e.g. "Worked 7-4").
    """
    strict = parse_time_range(text)
    if strict:
        return strict

    match = BARE_RANGE_RE.search(text)
    if not match:
        return None
    start = parse_time(match.group(1))
    end = parse_time(match.group(2))
    if not start or not end:
        return None
    end = make_after(end, start)
    duration = hours_between(start, end)
    if duration <= 0 or duration > 14:
        return None
    return {"start": start, "end": end, "duration": duration}


def _format_time(value):
    return value.strftime("%I:%M %p").lstrip("0") if value else ""


def parse_exception_text(description, schedule_info):
    """
    Interpret a free-text outside-schedule exception into paid hours.

    Patterns tried in order:
      1. Multi-segment schedule (4+ times, e.g. split shift like "8:30|1:00|2:00|6:00")
      2. Time range (two times) → schedule override for the day
      3. Single time + end-keyword ("late pickup X", "pay until X") → extend schedule end
      4. Single time + start-keyword ("came in at X", "started X") → extend schedule start
      5. Single time, no keyword → default to end-time extension (most common case)

    Returns:
      {"status": "parsed", "hours": float, "interpretation": str}
      {"status": "unparseable", "interpretation": str}
    """
    text = (description or "").strip()
    if not text:
        return {"status": "unparseable", "interpretation": ""}

    # Pattern 0: "pay hours worked" / "pay actual" / standalone "no break"
    # → pay the raw ADP hours as-is (no cap, no break adjustment)
    if re.search(
        r"pay\s+(?:the\s+)?(?:hours?\s+)?(?:worked|actual)|^\s*no\s+break\s*$",
        text,
        re.I,
    ) or re.search(r"pay\s+hours\s+worked", text, re.I):
        return {
            "status": "pay_actual",
            "interpretation": "Pay actual hours worked (no schedule cap, no break)",
        }

    tokens = TIME_TOKEN_RE.findall(text)
    parsed_times = [parsed for parsed in (parse_time(tok) for tok in tokens) if parsed]

    # Pattern 1: multi-segment / split shift (4+ paired times)
    # Gaps between segments are unpaid by construction — don't apply a break deduction.
    if len(parsed_times) >= 4 and len(parsed_times) % 2 == 0:
        total = 0.0
        segments = []
        for i in range(0, len(parsed_times), 2):
            seg_start = parsed_times[i]
            seg_end = make_after(parsed_times[i + 1], seg_start)
            span = hours_between(seg_start, seg_end)
            if span <= 0 or span > 14:
                total = -1
                break
            total += span
            segments.append((seg_start, seg_end))
        if total > 0 and total <= 14:
            total = round(total, 2)
            segs_str = ", ".join(
                f"{_format_time(s)}–{_format_time(e)}" for s, e in segments
            )
            return {
                "status": "parsed",
                "hours": total,
                "interpretation": f"Split-shift {segs_str} = {total:.2f} hrs",
            }

    # Pattern 2: two times → schedule override
    timerange = _extract_time_range(text)
    if timerange:
        break_hours = schedule_info.get("break_hours", 0.0) if schedule_info else 0.0
        hours = round(max(0.0, timerange["duration"] - break_hours), 2)
        if 0 < hours <= 14:
            brk = f" − {break_hours:g}h break" if break_hours else ""
            return {
                "status": "parsed",
                "hours": hours,
                "interpretation": (
                    f"Schedule override {_format_time(timerange['start'])}"
                    f"–{_format_time(timerange['end'])}{brk} = {hours:.2f} hrs"
                ),
            }

    # Single time + schedule → start-shift or end-extension
    if schedule_info:
        times = parsed_times

        if len(times) == 1:
            sched_start = schedule_info.get("start")
            sched_end = schedule_info.get("end")
            break_hours = schedule_info.get("break_hours", 0.0)
            clock = times[0]

            is_start = bool(START_TIME_PHRASES.search(text))
            is_end = bool(END_TIME_PHRASES.search(text))

            # Explicit start-shift keyword: treat as new start time
            if is_start and not is_end and sched_end:
                new_start = clock
                duration = hours_between(new_start, sched_end)
                if duration > 0:
                    hours = round(max(0.0, duration - break_hours), 2)
                    if 0 < hours <= 14:
                        return {
                            "status": "parsed",
                            "hours": hours,
                            "interpretation": (
                                f"Started {_format_time(new_start)} (early) = {hours:.2f} hrs"
                            ),
                        }

            # Default / end-keyword: extend end of shift to this time
            if sched_start:
                new_end = make_after(clock, sched_start)
                duration = hours_between(sched_start, new_end)
                if duration > 0:
                    hours = round(max(0.0, duration - break_hours), 2)
                    if 0 < hours <= 14:
                        verb = "Extended end to" if is_end else "End at"
                        return {
                            "status": "parsed",
                            "hours": hours,
                            "interpretation": f"{verb} {_format_time(new_end)} = {hours:.2f} hrs",
                        }

    return {
        "status": "unparseable",
        "interpretation": f"Could not interpret: {text}",
    }


# ---------------------------------------------------------------------------
# PUNCH CORRECTION PARSING
# ---------------------------------------------------------------------------

def parse_punch_correction(text):
    raw = cell_to_text(text)
    if not raw:
        return {"status": "review", "reason": "blank correction"}

    token_texts = TIME_TOKEN_RE.findall(raw)
    token_times = [parse_time(token) for token in token_texts]

    if "break" in raw.lower() and len(token_times) >= 4 and all(token_times[:4]):
        start = token_times[0]
        end = make_after(token_times[1], start)
        break_start = make_after(token_times[2], start)
        break_end = make_after(token_times[3], break_start)
        total = round(
            hours_between(start, end) - hours_between(break_start, break_end),
            2,
        )
        if total <= 0 or total > 12:
            return {"status": "review", "reason": f"implausible total correction hours: {raw}"}
        return {
            "status": "parsed",
            "hours": total,
            "start_time": start,
            "summary": raw,
        }

    if "|" in raw:
        parts = [piece.strip() for piece in raw.split("|") if piece.strip()]
        if len(parts) % 2 != 0:
            return {"status": "review", "reason": f"odd number of punch times: {raw}"}

        times = [parse_time(part) for part in parts]
        if any(t is None for t in times):
            return {"status": "review", "reason": f"could not parse punch times: {raw}"}

        total = 0.0
        for index in range(0, len(times), 2):
            start = times[index]
            end = make_after(times[index + 1], start)
            span = hours_between(start, end)
            if span <= 0 or span > 12:
                return {"status": "review", "reason": f"implausible segment in correction: {raw}"}
            total += span

        total = round(total, 2)
        if total > 12:
            return {"status": "review", "reason": f"implausible total correction hours: {raw}"}

        return {
            "status": "parsed",
            "hours": total,
            "start_time": times[0],
            "summary": raw,
        }

    timerange = parse_time_range(raw)
    if not timerange:
        return {"status": "review", "reason": f"could not parse correction: {raw}"}

    break_hours = parse_break_hours(raw)
    total = round(max(0.0, timerange["duration"] - break_hours), 2)
    if total <= 0 or total > 12:
        return {"status": "review", "reason": f"implausible total correction hours: {raw}"}

    return {
        "status": "parsed",
        "hours": total,
        "start_time": timerange["start"],
        "summary": raw,
    }


# ---------------------------------------------------------------------------
# INPUT PARSING
# ---------------------------------------------------------------------------

def parse_adp_csv(path):
    entries = []
    pay_period_start = None
    pay_period_end = None
    current_name = None

    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for row_index, row in enumerate(reader):
            if row_index == 0 and len(row) >= 3 and row[0].strip() == "Date range":
                pay_period_start = parse_date_value(row[1].strip())
                pay_period_end = parse_date_value(row[2].strip())
                continue

            if len(row) < 6:
                continue

            first = row[0].strip()
            if first in ("Employee Name", "") and not current_name:
                continue

            if first:
                current_name = first.strip('"')

            if not current_name:
                continue

            work_date = parse_date_value(row[2].strip())
            if not work_date:
                continue

            hours = parse_float(row[5]) or 0.0
            in_time_text = row[3].strip()
            out_time_text = row[4].strip()

            entries.append(
                {
                    "employee": current_name,
                    "date": work_date,
                    "in_time": in_time_text,
                    "out_time": out_time_text,
                    "in_time_obj": parse_time(in_time_text),
                    "out_time_obj": parse_time(out_time_text),
                    "hours": hours,
                    "note": row[9].strip() if len(row) > 9 else "",
                }
            )

    return {
        "entries": entries,
        "pay_period_start": pay_period_start,
        "pay_period_end": pay_period_end,
    }


def parse_adp_xlsx(path):
    """Parse an ADP export saved as Excel (.xlsx). Same column layout as the CSV."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active

    entries = []
    pay_period_start = None
    pay_period_end = None
    current_name = None

    for row in worksheet.iter_rows(values_only=True):
        # Row 1: Date range header
        if row[0] == "Date range" and len(row) >= 3:
            pay_period_start = parse_date_value(row[1])
            pay_period_end = parse_date_value(row[2])
            continue

        if row[0] == "Employee Name":
            continue

        if not any(row):
            continue

        name = cell_to_text(row[0])
        if name:
            current_name = name

        if not current_name:
            continue

        work_date = parse_date_value(row[2])
        if not work_date:
            continue

        in_time_obj = parse_time(row[3])
        out_time_obj = parse_time(row[4])
        # Format times as strings for display
        in_time_str = in_time_obj.strftime("%I:%M %p").lstrip("0") if in_time_obj else ""
        out_time_str = out_time_obj.strftime("%I:%M %p").lstrip("0") if out_time_obj else ""

        hours = parse_float(row[5]) or 0.0
        note = cell_to_text(row[9]) if len(row) > 9 else ""

        entries.append(
            {
                "employee": current_name,
                "date": work_date,
                "in_time": in_time_str,
                "out_time": out_time_str,
                "in_time_obj": in_time_obj,
                "out_time_obj": out_time_obj,
                "hours": hours,
                "note": note,
            }
        )

    return {
        "entries": entries,
        "pay_period_start": pay_period_start,
        "pay_period_end": pay_period_end,
    }


def parse_adp(path):
    """Dispatch to the correct ADP parser based on file extension."""
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return parse_adp_xlsx(path)
    return parse_adp_csv(path)


def _extract_kv(cells):
    data = {}
    index = 0
    while index < len(cells):
        cell = cells[index]
        if cell == "Date:" and index + 1 < len(cells):
            data["date_str"] = cells[index + 1]
            index += 2
        elif cell == "Employee name:" and index + 1 < len(cells):
            data["employee"] = cells[index + 1]
            index += 2
        elif cell == "Time in:" and index + 1 < len(cells):
            data["time_in_str"] = cells[index + 1]
            index += 2
        elif cell == "Time out:" and index + 1 < len(cells):
            data["time_out_str"] = cells[index + 1]
            index += 2
        elif cell == "Why needed?" and index + 1 < len(cells):
            data["reason"] = cells[index + 1]
            index += 2
        else:
            index += 1
    return data


def parse_docx_notes(path):
    notes = {
        "source_type": "docx",
        "employees": {},
        "missed_punches": [],
        "outside_schedule": [],
        "leave_entries": [],
        "review_items": [],
        "prompts": [],
    }

    document = docx.Document(path)
    for table in document.tables:
        for row in table.rows:
            cells = [cell.text.strip() for cell in row.cells]
            deduped = []
            for value in cells:
                if not deduped or value != deduped[-1]:
                    deduped.append(value)

            first = deduped[0] if deduped else ""
            data = _extract_kv(deduped)

            if "MISSED CLOCK-IN" in first.upper():
                work_date = parse_date_value(data.get("date_str"))
                employee = data.get("employee", "").strip()
                note_text = data.get("time_out_str", "").strip()
                if not work_date or not employee:
                    continue

                explicit = parse_punch_correction(note_text)
                if explicit["status"] == "parsed":
                    text = explicit["summary"]
                else:
                    missed_match = re.search(
                        r"missed punch at\s+(\d{1,2}:\d{2}(?:\s*[ap]m)?)",
                        note_text,
                        re.I,
                    )
                    time_in = data.get("time_in_str", "").strip()
                    if missed_match and time_in:
                        text = f"{time_in}-{missed_match.group(1)}"
                    else:
                        text = note_text

                notes["missed_punches"].append(
                    {
                        "employee": employee,
                        "date": work_date,
                        "text": text,
                        "source_note": note_text,
                    }
                )

            if "APPROVED TO WORK OUTSIDE" in first.upper():
                work_dates = expand_date_expression(data.get("date_str"))
                employee = data.get("employee", "").strip()
                reason = data.get("reason", "").strip()
                for work_date in work_dates:
                    notes["outside_schedule"].append(
                        {
                            "employee": employee,
                            "date": work_date,
                            "description": reason,
                            "schedule_override": None,
                        }
                    )

    return notes


def pair_date_lines_with_descriptions(date_value, description_value):
    date_lines = split_nonempty_lines(date_value)
    description_lines = split_nonempty_lines(description_value)

    if not date_lines:
        return []

    if len(date_lines) == len(description_lines) and len(date_lines) > 1:
        return list(zip(date_lines, description_lines))

    description = "\n".join(description_lines).strip()
    return [(date_line, description) for date_line in date_lines]


def parse_workbook_prompts(worksheet):
    prompts = []
    prompt_section = False
    for row in range(1, worksheet.max_row + 1):
        first = cell_to_text(worksheet.cell(row, 1).value)
        if first == "Prompts:":
            prompt_section = True
            continue
        if not prompt_section:
            continue
        if first:
            prompts.append(first)
    return prompts


def parse_xlsx_notes(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    # Find the header row dynamically — it's the row whose col 1 = "Employee".
    # This tolerates blank rows at the top that vary file to file.
    header_row = None
    for r in range(1, min(20, worksheet.max_row + 1)):
        if cell_to_text(worksheet.cell(r, 1).value).strip() == "Employee":
            header_row = r
            break
    if header_row is None:
        header_row = 2  # fallback

    # Detect whether the sheet has a dedicated "Break Time" column (col 4).
    # New format: Employee | Status | Schedule | Break Time | Missed date | ...
    # Old format: Employee | Status | Schedule | Missed date | ...
    col4_header = cell_to_text(worksheet.cell(header_row, 4).value).strip().lower()
    has_break_col = "break" in col4_header
    # Column offset: 1 when Break Time column is present, 0 otherwise
    o = 1 if has_break_col else 0
    # Data rows start two rows after the header (one sub-header row in between)
    data_start = header_row + 2

    notes = {
        "source_type": "xlsx",
        "employees": {},
        "missed_punches": [],
        "outside_schedule": [],
        "leave_entries": [],
        "review_items": [],
        "prompts": parse_workbook_prompts(worksheet),
    }

    for row in range(data_start, worksheet.max_row + 1):
        employee = cell_to_text(worksheet.cell(row, 1).value)
        if employee == "Additional Notes:":
            break
        if not employee:
            continue

        status = cell_to_text(worksheet.cell(row, 2).value)
        schedule_text = cell_to_text(worksheet.cell(row, 3).value)

        # Break hours: from dedicated column when present, else parsed from schedule text
        if has_break_col:
            break_hours = parse_break_hours(cell_to_text(worksheet.cell(row, 4).value))
        else:
            break_hours = parse_break_hours(schedule_text)

        schedule = parse_schedule_text(schedule_text, override_break=break_hours if has_break_col else None)
        notes["employees"][employee] = {
            "status": status,
            "schedule_text": schedule_text,
            "schedule": schedule,
        }

        # Pair each missed-punch date with its corresponding punch sequence.
        # Handles three layouts Megha uses:
        #   1. Single date + single punch line  (most common)
        #   2. Multiple dates on separate lines + matching punch lines (newline-separated)
        #   3. Multiple dates on one line (space-separated) + matching punch lines
        missed_date_raw = worksheet.cell(row, 4 + o).value
        missed_text_raw = cell_to_text(worksheet.cell(row, 5 + o).value)

        all_dates = expand_date_expression(missed_date_raw)
        text_lines = [ln.strip() for ln in missed_text_raw.splitlines() if ln.strip()]

        if len(all_dates) > 1 and len(all_dates) == len(text_lines):
            # Each date pairs with the same-position text line
            date_text_pairs = list(zip(all_dates, text_lines))
        else:
            # Single date or counts don't match — apply full text to every date
            date_text_pairs = [(d, missed_text_raw) for d in all_dates]

        for work_date, missed_text in date_text_pairs:
            if not missed_text:
                continue
            notes["missed_punches"].append(
                {
                    "employee": employee,
                    "date": work_date,
                    "text": missed_text,
                    "source_note": missed_text,
                }
            )

        outside_dates_raw = worksheet.cell(row, 6 + o).value
        outside_hours_raw = worksheet.cell(row, 7 + o).value
        all_outside_dates = expand_date_expression(outside_dates_raw)
        if all_outside_dates:
            hours_lines = [ln.strip() for ln in cell_to_text(outside_hours_raw).splitlines() if ln.strip()]
            whole_text = cell_to_text(outside_hours_raw).strip()

            if len(hours_lines) == len(all_outside_dates):
                # Per-date pairing: each date gets its own line as description + numeric attempt
                date_entries = [
                    (d, line, parse_float(line))
                    for d, line in zip(all_outside_dates, hours_lines)
                ]
            else:
                # Counts don't match — fall back to whole-cell text for each date.
                # Divide numeric total evenly if possible, else leave as None for text parsing.
                total_hrs = parse_float(outside_hours_raw)
                per_day = round(total_hrs / len(all_outside_dates), 4) if total_hrs else None
                date_entries = [(d, whole_text, per_day) for d in all_outside_dates]

            for work_date, description, approved_hours in date_entries:
                notes["outside_schedule"].append(
                    {
                        "employee": employee,
                        "date": work_date,
                        "description": description,
                        "approved_hours": approved_hours,
                    }
                )

        # Time Off: dates are no longer required — just the hours total as a lump sum
        time_off_hours = parse_float(worksheet.cell(row, 9 + o).value)
        if time_off_hours:
            notes["leave_entries"].append(
                {
                    "kind": "Time Off",
                    "employee": employee,
                    "dates": [],
                    "date_label": "",
                    "hours": round(time_off_hours, 2),
                }
            )

        holiday_dates = worksheet.cell(row, 10 + o).value
        holiday_hours = parse_float(worksheet.cell(row, 11 + o).value)
        if expand_date_expression(holiday_dates) and holiday_hours:
            notes["leave_entries"].append(
                {
                    "kind": "Holiday",
                    "employee": employee,
                    "dates": expand_date_expression(holiday_dates),
                    "date_label": cell_to_text(holiday_dates),
                    "hours": round(holiday_hours, 2),
                }
            )

    return notes


def parse_notes(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".docx":
        return parse_docx_notes(path)
    if suffix in (".xlsx", ".xlsm"):
        return parse_xlsx_notes(path)
    raise ValueError(f"Unsupported notes file: {path}")


def parse_workbook_rules(notes):
    rules = {
        "cap_to_schedule": notes.get("source_type") == "xlsx",
        "add_leave": notes.get("source_type") == "xlsx",
        "special_full_day": None,
    }

    for prompt in notes.get("prompts", []):
        lowered = prompt.lower()
        if "calculate based on scheduled hours" in lowered:
            rules["cap_to_schedule"] = True
        if "add holiday" in lowered or "time off" in lowered:
            rules["add_leave"] = True

        match = re.search(
            r"pay full day\s*\((\d+(?:\.\d+)?)\s*hours\)\s*for\s*(\d{1,2}/\d{1,2}/\d{2,4})",
            prompt,
            re.I,
        )
        if match:
            rules["special_full_day"] = {
                "hours": float(match.group(1)),
                "date": parse_date_value(match.group(2)),
                "skip_part_time": "part-time" in lowered,
                "skip_if_time_off": "time off used" in lowered,
                "deduct_lateness": "deduct for lateness" in lowered,
            }

    return rules


# ---------------------------------------------------------------------------
# CALCULATION
# ---------------------------------------------------------------------------

def initial_day_record(employee, work_date):
    return {
        "employee": employee,
        "date": work_date,
        "raw_entries": [],
        "raw_hours": 0.0,
        "raw_first_in": None,
        "raw_last_out": None,
        "corrected_hours": None,
        "corrected_text": "",
        "corrected_start": None,
        "actual_hours": 0.0,
        "final_work_hours": 0.0,
        "schedule_hours": None,
        "schedule_label": "",
        "outside_description": "",
        "outside_approved": False,
        "special_adjustment": None,
        "anomalies": [],
        "notes": [],
    }


def build_day_records(entries):
    records = {}
    for entry in entries:
        key = (entry["employee"], entry["date"])
        if key not in records:
            records[key] = initial_day_record(entry["employee"], entry["date"])
        record = records[key]
        record["raw_entries"].append(entry)
        record["raw_hours"] = round(record["raw_hours"] + entry["hours"], 2)

        in_time = entry.get("in_time_obj")
        if in_time and (record["raw_first_in"] is None or in_time < record["raw_first_in"]):
            record["raw_first_in"] = in_time

        out_time = entry.get("out_time_obj")
        if out_time:
            adjusted_out = make_after(out_time, in_time) if in_time else out_time
            if record["raw_last_out"] is None or adjusted_out > record["raw_last_out"]:
                record["raw_last_out"] = adjusted_out
    return records


def get_or_create_day(records, employee, work_date):
    key = (employee, work_date)
    if key not in records:
        records[key] = initial_day_record(employee, work_date)
    return records[key]


def apply_notes(entries, notes, rules):
    employee_names = {entry["employee"] for entry in entries}
    employee_names.update(notes.get("employees", {}).keys())

    day_records = build_day_records(entries)
    employee_meta = notes.get("employees", {})

    logs = {
        "corrections": [],
        "leave": [],
        "exceptions": [],
        "anomalies": [],
        "review": list(notes.get("review_items", [])),
    }

    outside_map = {}
    for item in notes.get("outside_schedule", []):
        matched_name, error = resolve_employee_name(item["employee"], employee_names)
        if not matched_name:
            logs["review"].append(
                {
                    "employee": item["employee"],
                    "date": item["date"],
                    "issue": error,
                }
            )
            continue

        outside_map[(matched_name, item["date"])] = item
        logs["exceptions"].append(
            {
                "employee": matched_name,
                "date": item["date"],
                "description": item["description"],
            }
        )

    resolved_leave_entries = []
    leave_dates = defaultdict(lambda: {"Time Off": set(), "Holiday": set()})
    for item in notes.get("leave_entries", []):
        matched_name, error = resolve_employee_name(item["employee"], employee_names)
        if not matched_name:
            logs["review"].append(
                {
                    "employee": item["employee"],
                    "date": item["date_label"],
                    "issue": error,
                }
            )
            continue

        leave_dates[matched_name][item["kind"]].update(item["dates"])
        for work_date in item["dates"]:
            get_or_create_day(day_records, matched_name, work_date)
        log_item = {
            "employee": matched_name,
            "kind": item["kind"],
            "date_label": item["date_label"],
            "hours": round(item["hours"], 2),
            "allocated_hours": 0.0,
        }
        logs["leave"].append(log_item)
        resolved_leave_entries.append({**item, "employee": matched_name, "log_item": log_item})

    for item in notes.get("missed_punches", []):
        matched_name, error = resolve_employee_name(item["employee"], employee_names)
        if not matched_name:
            logs["corrections"].append(
                {
                    "employee": item["employee"],
                    "date": item["date"],
                    "status": "REVIEW",
                    "raw_hours": 0.0,
                    "corrected_hours": None,
                    "note": error,
                }
            )
            continue

        record = get_or_create_day(day_records, matched_name, item["date"])
        parsed = parse_punch_correction(item["text"])
        if parsed["status"] != "parsed":
            logs["corrections"].append(
                {
                    "employee": matched_name,
                    "date": item["date"],
                    "status": "REVIEW",
                    "raw_hours": record["raw_hours"],
                    "corrected_hours": None,
                    "note": parsed["reason"],
                }
            )
            record["notes"].append(f"Missed punch needs review: {parsed['reason']}")
            continue

        record["corrected_hours"] = parsed["hours"]
        record["corrected_text"] = parsed["summary"]
        record["corrected_start"] = parsed["start_time"]
        logs["corrections"].append(
            {
                "employee": matched_name,
                "date": item["date"],
                "status": "CORRECTED",
                "raw_hours": record["raw_hours"],
                "corrected_hours": parsed["hours"],
                "note": parsed["summary"],
            }
        )

    special_rule = rules.get("special_full_day")
    if special_rule and special_rule.get("date"):
        for employee, meta in employee_meta.items():
            status = cell_to_text(meta.get("status")).lower()
            if "full" not in status:
                continue
            if special_rule.get("skip_part_time") and "part-time" in status:
                continue
            get_or_create_day(day_records, employee, special_rule["date"])

    # Flag employees with punches but no schedule in notes — they can't be capped.
    flagged_no_schedule = set()
    for (employee, _), record in day_records.items():
        if employee in flagged_no_schedule:
            continue
        meta = employee_meta.get(employee)
        has_schedule = meta and meta.get("schedule", {}).get("mode") == "scheduled"
        has_punches = any(r["raw_hours"] > 0 for (e, _), r in day_records.items() if e == employee)
        if not has_schedule and has_punches:
            flagged_no_schedule.add(employee)
            logs["review"].append({
                "employee": employee,
                "date": "",
                "issue": "No schedule found in notes — hours not capped. Add this employee to the notes sheet or adjust manually.",
            })

    for key, record in day_records.items():
        employee = record["employee"]
        work_date = record["date"]
        meta = employee_meta.get(employee, {"status": "", "schedule_text": "", "schedule": {"mode": "actual"}})
        schedule = meta.get("schedule", {"mode": "actual"})
        outside = outside_map.get((employee, work_date))

        actual_hours = record["corrected_hours"] if record["corrected_hours"] is not None else record["raw_hours"]
        first_in = record["corrected_start"] or record["raw_first_in"]
        schedule_info = schedule_for_date(schedule, work_date)

        record["actual_hours"] = round(actual_hours, 2)
        record["outside_approved"] = bool(outside)
        record["outside_description"] = outside["description"] if outside else ""
        if schedule_info:
            record["schedule_hours"] = schedule_info["paid_hours"]
            record["schedule_label"] = schedule_info["label"]

        final_work = actual_hours
        approved_hours = outside.get("approved_hours") if outside else None

        # If no explicit number, try to interpret the description text
        # (e.g. "Late pickup 6:37", "Schedule 7 AM - 4 PM", "Worked 7-4")
        exception_parse = None
        if approved_hours is None and outside and outside.get("description"):
            exception_parse = parse_exception_text(outside["description"], schedule_info)

        if approved_hours is not None:
            # Megha wrote an explicit numeric hours value — pay that exactly
            final_work = approved_hours
            record["notes"].append(f"Approved: {approved_hours:.2f} hrs")
        elif exception_parse and exception_parse["status"] == "parsed":
            # Auto-interpreted a phrase like "Late pickup 6:37"
            final_work = exception_parse["hours"]
            record["notes"].append(
                f"Approved exception — {exception_parse['interpretation']} "
                f"(from: \"{outside['description']}\")"
            )
        elif exception_parse and exception_parse["status"] == "pay_actual":
            # "Pay hours worked" / "no break" — use raw ADP hours, no cap
            final_work = actual_hours
            record["notes"].append(
                f"Approved exception — {exception_parse['interpretation']} "
                f"(from: \"{outside['description']}\")"
            )
        else:
            if outside and outside.get("description"):
                # Description only and we couldn't parse it — informational note + review flag
                record["notes"].append(f"Note: {outside['description']}")
                if exception_parse and exception_parse["status"] == "unparseable":
                    logs["review"].append({
                        "employee": employee,
                        "date": work_date,
                        "issue": (
                            f"Exception note couldn't be auto-calculated "
                            f"(paid at schedule): \"{outside['description']}\""
                        ),
                    })
            if schedule_info and actual_hours > 0:
                # Cap paid hours to the schedule — overage only paid if Megha writes explicit approved hours.
                # Surface anomalies so she can see which days went over and decide.
                scheduled_hours = schedule_info["paid_hours"]
                sched_start = schedule_info.get("start")
                sched_end = schedule_info.get("end")
                last_out = record["raw_last_out"]

                TOLERANCE = 0.25  # 15 min
                WINDOW_TOLERANCE = timedelta(minutes=15)

                if actual_hours > scheduled_hours + TOLERANCE:
                    delta = round(actual_hours - scheduled_hours, 2)
                    desc = (
                        f"Worked {actual_hours:.2f} hrs — {delta:+.2f} over scheduled "
                        f"{scheduled_hours:.2f} (capped; add an exception to pay full)"
                    )
                    record["anomalies"].append(desc)
                    logs["anomalies"].append({
                        "employee": employee,
                        "date": work_date,
                        "type": "over_hours",
                        "actual_hours": round(actual_hours, 2),
                        "scheduled_hours": scheduled_hours,
                        "description": desc,
                    })

                if first_in and sched_start:
                    base = datetime(2000, 1, 1)
                    if datetime.combine(base, first_in) + WINDOW_TOLERANCE < datetime.combine(base, sched_start):
                        desc = (
                            f"Clocked in at {first_in.strftime('%I:%M %p').lstrip('0')} — "
                            f"before scheduled start {sched_start.strftime('%I:%M %p').lstrip('0')}"
                        )
                        record["anomalies"].append(desc)
                        logs["anomalies"].append({
                            "employee": employee,
                            "date": work_date,
                            "type": "early_in",
                            "description": desc,
                        })

                if last_out and sched_end:
                    base = datetime(2000, 1, 1)
                    if datetime.combine(base, last_out) > datetime.combine(base, sched_end) + WINDOW_TOLERANCE:
                        desc = (
                            f"Clocked out at {last_out.strftime('%I:%M %p').lstrip('0')} — "
                            f"after scheduled end {sched_end.strftime('%I:%M %p').lstrip('0')}"
                        )
                        record["anomalies"].append(desc)
                        logs["anomalies"].append({
                            "employee": employee,
                            "date": work_date,
                            "type": "late_out",
                            "description": desc,
                        })

                if actual_hours > scheduled_hours:
                    record["notes"].append(
                        f"Capped to schedule ({scheduled_hours:.2f} hrs)"
                    )
                final_work = min(actual_hours, scheduled_hours)

        if special_rule and work_date == special_rule.get("date"):
            status = cell_to_text(meta.get("status")).lower()
            has_time_off = work_date in leave_dates[employee]["Time Off"]
            if "full" in status and not (special_rule.get("skip_part_time") and "part-time" in status):
                if not (special_rule.get("skip_if_time_off") and has_time_off):
                    special_hours = special_rule["hours"]
                    if special_rule.get("deduct_lateness") and schedule_info and first_in:
                        if first_in > schedule_info["start"]:
                            late = hours_between(schedule_info["start"], first_in)
                            special_hours = max(0.0, special_hours - late)
                    special_hours = round(special_hours, 2)
                    if special_hours > final_work:
                        record["special_adjustment"] = special_hours
                        record["notes"].append(
                            f"Special full-day rule applied ({special_hours:.2f} hrs)"
                        )
                        final_work = special_hours

        record["final_work_hours"] = round(final_work, 2)

    leave_allocations = defaultdict(lambda: {"Time Off": 0.0, "Holiday": 0.0})

    for item in resolved_leave_entries:
        employee = item["employee"]
        remaining = round(item["hours"], 2)

        # Time Off with no dates: lump sum — add directly, no date-based spreading
        if item["kind"] == "Time Off" and not item["dates"]:
            leave_allocations[employee]["Time Off"] += remaining
            item["log_item"]["allocated_hours"] = remaining
            continue

        for work_date in sorted(item["dates"]):
            record = get_or_create_day(day_records, employee, work_date)
            target = record["schedule_hours"]

            if target is None:
                meta = employee_meta.get(employee, {})
                status = cell_to_text(meta.get("status")).lower()
                if (
                    special_rule
                    and work_date == special_rule.get("date")
                    and "full" in status
                ):
                    target = special_rule["hours"]

            # Holiday: pay the exact hours specified in the notes — never cap to schedule.
            # Time Off: cap to one day's scheduled hours per date, spread as needed.
            if item["kind"] == "Holiday":
                add_hours = remaining
            elif target is None:
                continue
            else:
                add_hours = round(min(remaining, target), 2)

            if add_hours <= 0:
                continue

            leave_allocations[employee][item["kind"]] += add_hours
            item["log_item"]["allocated_hours"] = round(
                item["log_item"]["allocated_hours"] + add_hours, 2
            )
            record["notes"].append(f"{item['kind']} allocated: {add_hours:.2f} hrs")
            remaining = round(remaining - add_hours, 2)
            if remaining <= 0.01:
                break

        if remaining > 0.01 and item["dates"]:
            logs["review"].append(
                {
                    "employee": employee,
                    "date": item["date_label"],
                    "issue": (
                        f"Unallocated {item['kind'].lower()} hours remaining "
                        f"({remaining:.2f}); split hours by date in the notes sheet"
                    ),
                }
            )

    return day_records, logs, leave_allocations


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
FLAG_FILL = PatternFill("solid", fgColor="FFE699")
ERROR_FILL = PatternFill("solid", fgColor="FF9999")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(worksheet, row, col, value, fill=None, font=None):
    cell = worksheet.cell(row=row, column=col, value=value)
    cell.fill = fill or HEADER_FILL
    cell.font = font or HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def write_cell(worksheet, row, col, value, fill=None, bold=False, fmt=None, wrap=False):
    cell = worksheet.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = BORDER
    return cell


def write_excel(output_path, day_records, logs, leave_totals, employee_meta, pay_dates, pay_period_start, pay_period_end):
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"

    all_employees = sorted(
        set(employee_meta.keys()) | {record["employee"] for record in day_records.values()} | set(leave_totals.keys()),
        key=normalize_name,
    )

    title = f"Payroll Hours Summary  |  Pay Period: {pay_period_start.strftime('%m/%d/%Y')} – {pay_period_end.strftime('%m/%d/%Y')}"
    last_col = 7 + len(pay_dates)
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = summary.cell(row=1, column=1, value=title)
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    columns = [
        "Employee",
        "Status",
        "Schedule",
        "Worked Hrs",
        "Time Off",
        "Holiday",
        "Total Payable",
    ]
    for index, value in enumerate(columns, start=1):
        hdr(summary, 2, index, value)

    for index, work_date in enumerate(pay_dates, start=8):
        hdr(summary, 2, index, work_date.strftime("%a\n%m/%d"), fill=SUB_FILL, font=SUB_FONT)

    widths = [24, 14, 34, 12, 12, 12, 14]
    for index, width in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(index)].width = width
    for index in range(8, 8 + len(pay_dates)):
        summary.column_dimensions[get_column_letter(index)].width = 9

    records_by_employee = defaultdict(dict)
    for record in day_records.values():
        records_by_employee[record["employee"]][record["date"]] = record

    for row_index, employee in enumerate(all_employees, start=3):
        fill = ALT_FILL if row_index % 2 == 1 else None
        meta = employee_meta.get(employee, {})
        worked_total = round(
            sum(record["final_work_hours"] for record in records_by_employee.get(employee, {}).values()),
            2,
        )
        time_off = round(leave_totals[employee]["Time Off"], 2)
        holiday = round(leave_totals[employee]["Holiday"], 2)
        payable_total = round(worked_total + time_off + holiday, 2)

        write_cell(summary, row_index, 1, employee, fill=fill)
        write_cell(summary, row_index, 2, meta.get("status", ""), fill=fill)
        write_cell(summary, row_index, 3, meta.get("schedule_text", ""), fill=fill, wrap=True)
        write_cell(summary, row_index, 4, worked_total, fill=fill, fmt="0.00")
        write_cell(summary, row_index, 5, time_off, fill=fill, fmt="0.00")
        write_cell(summary, row_index, 6, holiday, fill=fill, fmt="0.00")
        write_cell(summary, row_index, 7, payable_total, fill=fill, fmt="0.00", bold=True)

        for offset, work_date in enumerate(pay_dates, start=8):
            record = records_by_employee.get(employee, {}).get(work_date)
            value = record["final_work_hours"] if record else ""
            write_cell(summary, row_index, offset, value, fill=fill, fmt="0.00")

    summary.freeze_panes = "A3"

    details = workbook.create_sheet("Daily Detail")
    detail_headers = [
        "Employee",
        "Date",
        "Day",
        "Raw Hours",
        "Correction",
        "Actual Hours",
        "Schedule Hrs",
        "Approved Exception",
        "Final Work Hrs",
        "Notes",
    ]
    detail_widths = [24, 12, 8, 10, 28, 12, 12, 18, 12, 60]
    for col, (header, width) in enumerate(zip(detail_headers, detail_widths), start=1):
        hdr(details, 1, col, header)
        details.column_dimensions[get_column_letter(col)].width = width

    for row_index, record in enumerate(
        sorted(day_records.values(), key=lambda item: (normalize_name(item["employee"]), item["date"])),
        start=2,
    ):
        has_anomaly = bool(record["anomalies"])
        has_correction = record["corrected_hours"] is not None or record["special_adjustment"]
        if has_anomaly:
            fill = ERROR_FILL
        elif has_correction:
            fill = FLAG_FILL
        else:
            fill = None
        combined_notes = list(record["notes"])
        for anomaly in record["anomalies"]:
            combined_notes.append(f"⚠ Anomaly: {anomaly}")
        write_cell(details, row_index, 1, record["employee"], fill=fill)
        write_cell(details, row_index, 2, record["date"].strftime("%m/%d/%Y"), fill=fill)
        write_cell(details, row_index, 3, record["date"].strftime("%a"), fill=fill)
        write_cell(details, row_index, 4, record["raw_hours"], fill=fill, fmt="0.00")
        write_cell(details, row_index, 5, record["corrected_text"], fill=fill, wrap=True)
        write_cell(details, row_index, 6, record["actual_hours"], fill=fill, fmt="0.00")
        write_cell(details, row_index, 7, record["schedule_hours"], fill=fill, fmt="0.00")
        write_cell(details, row_index, 8, "YES" if record["outside_approved"] else "", fill=fill)
        write_cell(details, row_index, 9, record["final_work_hours"], fill=fill, fmt="0.00")
        write_cell(details, row_index, 10, "\n".join(combined_notes), fill=fill, wrap=True)

    details.freeze_panes = "A2"

    review = workbook.create_sheet("Adjustments")
    for col, width in enumerate([24, 12, 12, 14, 60], start=1):
        review.column_dimensions[get_column_letter(col)].width = width

    row = 1
    hdr(review, row, 1, "Missed Punches / Corrections")
    review.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    for col, value in enumerate(["Employee", "Date", "Status", "Hours", "Note"], start=1):
        hdr(review, row, col, value, fill=SUB_FILL, font=SUB_FONT)
    row += 1
    for item in logs["corrections"]:
        fill = OK_FILL if item["status"] == "CORRECTED" else FLAG_FILL
        write_cell(review, row, 1, item["employee"], fill=fill)
        write_cell(review, row, 2, item["date"].strftime("%m/%d/%Y"), fill=fill)
        write_cell(review, row, 3, item["status"], fill=fill)
        hours_value = item["corrected_hours"] if item["corrected_hours"] is not None else item["raw_hours"]
        write_cell(review, row, 4, hours_value, fill=fill, fmt="0.00")
        write_cell(review, row, 5, item["note"], fill=fill, wrap=True)
        row += 1

    row += 1
    hdr(review, row, 1, "Leave Adjustments")
    review.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    for col, value in enumerate(["Employee", "Kind", "Date(s)", "Requested", "Allocated"], start=1):
        hdr(review, row, col, value, fill=SUB_FILL, font=SUB_FONT)
    row += 1
    for item in logs["leave"]:
        write_cell(review, row, 1, item["employee"], fill=ALT_FILL)
        write_cell(review, row, 2, item["kind"], fill=ALT_FILL)
        write_cell(review, row, 3, item["date_label"], fill=ALT_FILL, wrap=True)
        write_cell(review, row, 4, item["hours"], fill=ALT_FILL, fmt="0.00")
        write_cell(review, row, 5, item.get("allocated_hours", 0.0), fill=ALT_FILL, fmt="0.00")
        row += 1

    row += 1
    hdr(review, row, 1, "Approved Schedule Exceptions")
    review.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    for col, value in enumerate(["Employee", "Date", "", "", "Description"], start=1):
        hdr(review, row, col, value, fill=SUB_FILL, font=SUB_FONT)
    row += 1
    for item in logs["exceptions"]:
        write_cell(review, row, 1, item["employee"], fill=ALT_FILL)
        write_cell(review, row, 2, item["date"].strftime("%m/%d/%Y"), fill=ALT_FILL)
        write_cell(review, row, 5, item["description"], fill=ALT_FILL, wrap=True)
        row += 1

    if logs.get("anomalies"):
        row += 1
        hdr(review, row, 1, "Schedule Anomalies (paid per labor law, flagged for review)")
        review.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for col, value in enumerate(["Employee", "Date", "Type", "", "Description"], start=1):
            hdr(review, row, col, value, fill=SUB_FILL, font=SUB_FONT)
        row += 1
        for item in logs["anomalies"]:
            write_cell(review, row, 1, item["employee"], fill=FLAG_FILL)
            write_cell(review, row, 2, item["date"].strftime("%m/%d/%Y"), fill=FLAG_FILL)
            write_cell(review, row, 3, item["type"], fill=FLAG_FILL)
            write_cell(review, row, 5, item["description"], fill=FLAG_FILL, wrap=True)
            row += 1

    if logs["review"]:
        row += 1
        hdr(review, row, 1, "Needs Review")
        review.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for col, value in enumerate(["Employee", "Date", "", "", "Issue"], start=1):
            hdr(review, row, col, value, fill=SUB_FILL, font=SUB_FONT)
        row += 1
        for item in logs["review"]:
            write_cell(review, row, 1, item.get("employee", ""), fill=ERROR_FILL)
            write_cell(review, row, 2, cell_to_text(item.get("date", "")), fill=ERROR_FILL)
            write_cell(review, row, 5, item.get("issue", ""), fill=ERROR_FILL, wrap=True)
            row += 1

    raw = workbook.create_sheet("Raw Punches")
    raw_headers = ["Employee", "Date", "In Time", "Out Time", "Hours", "Note"]
    for col, header in enumerate(raw_headers, start=1):
        hdr(raw, 1, col, header)
        raw.column_dimensions[get_column_letter(col)].width = [24, 12, 12, 12, 10, 40][col - 1]

    raw_row = 2
    for record in sorted(day_records.values(), key=lambda item: (normalize_name(item["employee"]), item["date"])):
        for entry in record["raw_entries"]:
            write_cell(raw, raw_row, 1, entry["employee"])
            write_cell(raw, raw_row, 2, entry["date"].strftime("%m/%d/%Y"))
            write_cell(raw, raw_row, 3, entry["in_time"])
            write_cell(raw, raw_row, 4, entry["out_time"])
            write_cell(raw, raw_row, 5, entry["hours"], fmt="0.00")
            write_cell(raw, raw_row, 6, entry["note"], wrap=True)
            raw_row += 1

    workbook.save(output_path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 3:
        print("Usage: python3 payroll_calculator.py <ADPPayroll.csv> <notes.docx|notes.xlsx>")
        sys.exit(1)

    csv_path = Path(sys.argv[1]).expanduser()
    notes_path = Path(sys.argv[2]).expanduser()
    output_path = csv_path.parent / f"{csv_path.stem}_corrected.xlsx"

    print(f"Reading ADP data:      {csv_path}")
    adp = parse_adp(csv_path)
    entries = adp["entries"]
    print(f"  -> {len(entries)} punch rows across {len(set(entry['employee'] for entry in entries))} employees")

    print(f"Reading payroll notes: {notes_path}")
    notes = parse_notes(notes_path)
    print(
        "  -> "
        f"{len(notes.get('employees', {}))} employee note rows, "
        f"{len(notes.get('missed_punches', []))} missed punch entries, "
        f"{len(notes.get('outside_schedule', []))} schedule exceptions, "
        f"{len(notes.get('leave_entries', []))} leave adjustments"
    )

    rules = parse_workbook_rules(notes)
    day_records, logs, leave_totals = apply_notes(entries, notes, rules)

    pay_period_start = adp["pay_period_start"]
    pay_period_end = adp["pay_period_end"]
    if not pay_period_start or not pay_period_end:
        dates = sorted(record["date"] for record in day_records.values())
        pay_period_start = dates[0]
        pay_period_end = dates[-1]

    pay_dates = weekday_dates_only(pay_period_start, pay_period_end)

    print("\nCalculated totals:")
    all_employees = sorted(
        set(notes.get("employees", {}).keys()) | {record["employee"] for record in day_records.values()} | set(leave_totals.keys()),
        key=normalize_name,
    )
    for employee in all_employees:
        worked = round(
            sum(record["final_work_hours"] for record in day_records.values() if record["employee"] == employee),
            2,
        )
        time_off = round(leave_totals[employee]["Time Off"], 2)
        holiday = round(leave_totals[employee]["Holiday"], 2)
        payable = round(worked + time_off + holiday, 2)
        print(
            f"  {employee:<24} worked={worked:>6.2f}  "
            f"time_off={time_off:>6.2f}  holiday={holiday:>6.2f}  total={payable:>6.2f}"
        )

    if logs["corrections"]:
        print("\nCorrections:")
        for item in logs["corrections"]:
            date_text = item["date"].strftime("%m/%d/%Y")
            if item["status"] == "CORRECTED":
                print(
                    f"  [OK]     {item['employee']} {date_text}: "
                    f"{item['raw_hours']:.2f} -> {item['corrected_hours']:.2f}"
                )
            else:
                print(f"  [REVIEW] {item['employee']} {date_text}: {item['note']}")

    if logs["review"]:
        print("\nNeeds review:")
        for item in logs["review"]:
            print(f"  - {item.get('employee', '')} {cell_to_text(item.get('date', ''))}: {item.get('issue', '')}")

    write_excel(
        output_path=output_path,
        day_records=day_records,
        logs=logs,
        leave_totals=leave_totals,
        employee_meta=notes.get("employees", {}),
        pay_dates=pay_dates,
        pay_period_start=pay_period_start,
        pay_period_end=pay_period_end,
    )
    print(f"\nReport saved to: {output_path}")


if __name__ == "__main__":
    main()
